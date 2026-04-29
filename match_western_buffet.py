#!/usr/bin/env python3
"""
Match dishes from "Corporate -Western Buffet (New)" sheet in the LaCasa spreadsheet
against master_dishes.xml product templates.

Output: structured list of ALL dishes with category, names, surcharge, and match status.
"""

from __future__ import annotations
import re
import xml.etree.ElementTree as ET
import openpyxl


# ── 1. Parse master_dishes.xml ──────────────────────────────────────────────

def parse_master_dishes(xml_path: str) -> list[dict]:
    """Return list of {id, name_en, name_cn, full_name} from XML."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    dishes = []
    for rec in root.findall("record"):
        if rec.get("model") != "product.template":
            continue
        xml_id = rec.get("id")
        name_field = rec.find("field[@name='name']")
        if name_field is None or not name_field.text:
            continue
        full = name_field.text.strip()
        # Name format: "English name\nChinese name"
        lines = full.split("\n")
        name_en = lines[0].strip()
        name_cn = lines[1].strip() if len(lines) > 1 else ""
        dishes.append({
            "id": xml_id,
            "name_en": name_en,
            "name_cn": name_cn,
            "full_name": full,
        })
    return dishes


# ── 2. Read spreadsheet ────────────────────────────────────────────────────

def read_western_buffet(xlsx_path: str) -> list[dict]:
    """
    Read all rows. Return list of dicts with keys:
      row_num, name_en, name_cn, is_header, category, surcharge
    Section headers are detected by pattern like "A. ...", "B. ...", etc.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Corporate -Western Buffet (New)"]

    # Header pattern: starts with a letter + "." like "A. Salad..."
    header_re = re.compile(r"^[A-Z]\.\s+")
    # Surcharge pattern: [+HK$XX per person]
    surcharge_re = re.compile(r'\[?\+\s*HK?\$(\d+)\s*(?:per\s*person)?\]?', re.IGNORECASE)

    rows = []
    current_category = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        # Column B = English, Column C = Chinese
        col_b = None
        col_c = None
        col_e = None  # cost column — dish rows usually have this
        for cell in row:
            if cell.column == 2:
                col_b = cell.value
            elif cell.column == 3:
                col_c = cell.value
            elif cell.column == 5:
                col_e = cell.value

        if not col_b or not isinstance(col_b, str):
            continue

        text = col_b.strip()
        if not text:
            continue

        # Skip the very first title row and non-food rows
        if "Western Lunch" in text or "Pricing" in text or "Price includes" in text:
            continue
        if "Food Price" in text or "Special charge" in text:
            continue
        if "Extra working" in text or "Remarks" in text or "Delivery charge" in text:
            continue
        if "Drinks Package" in text or "hours drinks" in text:
            continue
        if "Fruit Punch" in text or "House Wine" in text or "Corkage" in text:
            continue
        if "Add On Options" in text or "Plus $" in text or "Plus HK$" in text:
            continue
        if "Advance Order" in text or "working days" in text:
            continue
        if "Sales remarks" in text or "Sales offer" in text:
            continue
        if text.startswith("No drinks") or text.startswith("Disposable"):
            continue
        if text.startswith("Add canapes") or text.startswith("free") or text.startswith("discount"):
            continue
        if re.match(r"^\d+\.\s+(free|add|discount)", text):
            continue
        if "The price based" in text or "Standard Drink" in text or "Deluxe Drink" in text:
            continue
        if "Special offer" in text:
            continue

        # Extract surcharge
        surcharge = None
        sm = surcharge_re.search(text)
        if sm:
            surcharge = f"+${sm.group(1)}"

        # Check if this is a section header
        if header_re.match(text):
            current_category = text
            if col_c and isinstance(col_c, str):
                current_category = f"{text} / {col_c.strip()}"
            rows.append({
                "row_num": row[0].row,
                "name_en": text,
                "name_cn": col_c.strip() if col_c and isinstance(col_c, str) else "",
                "is_header": True,
                "category": current_category,
                "surcharge": None,
            })
            continue

        # It's a dish row if column B has text and it's not a header/metadata
        cn = col_c.strip() if col_c and isinstance(col_c, str) else ""
        rows.append({
            "row_num": row[0].row,
            "name_en": text,
            "name_cn": cn,
            "is_header": False,
            "category": current_category,
            "surcharge": surcharge,
        })

    return rows


# ── 3. Matching logic ──────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Lowercase, strip parenthetical markers like (v), (V), collapse whitespace."""
    s = s.lower().strip()
    # Remove (v), (素), piece counts like (40 pcs), price surcharges like [+HK$...]
    s = re.sub(r"\(v\)", "", s)
    s = re.sub(r"\(素\)", "", s)
    s = re.sub(r"\(\d+\s*pcs?\)", "", s)
    s = re.sub(r"\(\d+\s*件\)", "", s)
    s = re.sub(r"\(\d+\s*杯\)", "", s)
    s = re.sub(r"\[.*?\]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def get_significant_words(s: str, min_words: int = 3) -> list[str]:
    """Extract significant words (skip short/common ones)."""
    stop = {"with", "and", "the", "a", "an", "of", "in", "on", "for", "&", "served", "w/"}
    words = normalize(s).split()
    sig = [w for w in words if w not in stop and len(w) > 1]
    return sig[:max(min_words, len(sig))]


def find_match(dish_en: str, master_dishes: list[dict]) -> tuple[str | None, str]:
    """
    Try to match a buffet dish name against master dishes.
    Returns (xml_id or None, match_type).
    """
    norm_dish = normalize(dish_en)

    # Pass 1: Exact match (normalized)
    for m in master_dishes:
        if normalize(m["name_en"]) == norm_dish:
            return m["id"], "EXACT"

    # Pass 2: Contains match (one contains the other)
    # Require the shorter string to be at least 20 chars to avoid false positives
    for m in master_dishes:
        nm = normalize(m["name_en"])
        shorter = min(len(norm_dish), len(nm))
        if shorter >= 20 and (norm_dish in nm or nm in norm_dish):
            return m["id"], "CONTAINS"

    # Pass 3: Significant words overlap (>= 75% of buffet dish words found in master)
    dish_words = set(get_significant_words(dish_en))
    distinguishing = {"vegan", "charcoal", "angus", "halibut", "grouper", "lamb", "pork", "chicken", "beef"}
    dish_dist = dish_words & distinguishing
    if len(dish_words) >= 2:
        best_match = None
        best_score = 0
        for m in master_dishes:
            master_words = set(get_significant_words(m["name_en"]))
            master_dist = master_words & distinguishing
            # If distinguishing words differ, skip
            if dish_dist != master_dist:
                continue
            overlap = dish_words & master_words
            if len(dish_words) > 0:
                score = len(overlap) / len(dish_words)
                if score > best_score:
                    best_score = score
                    best_match = m
        if best_score >= 0.75 and best_match:
            return best_match["id"], f"WORDS({best_score:.0%})"

    return None, "NONE"


# ── 4. Main ────────────────────────────────────────────────────────────────

def main():
    xml_path = "/Users/felixyuen/Desktop/LCS/lcs_product_catalog/data/master_dishes.xml"
    xlsx_path = "/Users/felixyuen/Downloads/LaCasa-alacart_2025.xlsx"

    master = parse_master_dishes(xml_path)
    print(f"Master dishes loaded: {len(master)} product templates\n")

    rows = read_western_buffet(xlsx_path)

    # ── Section 1: Category headers with choose-X rules ──
    print("=" * 90)
    print("CATEGORY SECTION HEADERS & CHOOSE-X RULES")
    print("=" * 90)
    for row in rows:
        if row["is_header"]:
            print(f"  EN: {row['name_en']}")
            print(f"  CN: {row['name_cn']}")
            print()

    # ── Section 2: Complete structured dish list ──
    print("=" * 90)
    print("COMPLETE DISH LIST — ALL DISHES")
    print("=" * 90)

    dish_num = 0
    matched_count = 0
    unmatched_count = 0
    current_cat = None

    for row in rows:
        if row["is_header"]:
            current_cat = row["category"]
            print(f"\n{'─' * 90}")
            print(f"  CATEGORY: {row['name_en']}")
            if row["name_cn"]:
                print(f"            {row['name_cn']}")
            print(f"{'─' * 90}")
            continue

        dish_num += 1
        xml_id, match_type = find_match(row["name_en"], master)

        if xml_id:
            matched_count += 1
            master_dish = next(m for m in master if m["id"] == xml_id)
            match_label = xml_id
        else:
            unmatched_count += 1
            match_label = "NEW"

        surcharge_str = f"  Surcharge: {row['surcharge']}" if row["surcharge"] else ""

        print(f"\n  [{dish_num:02d}] Match: {match_label}  ({match_type}){surcharge_str}")
        print(f"       EN: {row['name_en']}")
        print(f"       CN: {row['name_cn']}")
        if xml_id:
            print(f"       Master EN: {master_dish['name_en']}")

    # ── Section 3: Summary ──
    print(f"\n\n{'=' * 90}")
    print(f"SUMMARY")
    print(f"{'=' * 90}")
    print(f"  Total dishes:   {dish_num}")
    print(f"  Matched:        {matched_count}")
    print(f"  Unmatched:      {unmatched_count}  (need NEW products)")
    print()

    # ── Section 4: Unmatched list ──
    if unmatched_count:
        print(f"{'=' * 90}")
        print("UNMATCHED DISHES — Need New XML IDs")
        print(f"{'=' * 90}")
        for row in rows:
            if row["is_header"]:
                continue
            xml_id, _ = find_match(row["name_en"], master)
            if not xml_id:
                surcharge_str = f"  (Surcharge: {row['surcharge']})" if row["surcharge"] else ""
                print(f"  EN: {row['name_en']}{surcharge_str}")
                print(f"  CN: {row['name_cn']}")
                print(f"  Category: {row['category']}")
                print()


if __name__ == "__main__":
    main()
